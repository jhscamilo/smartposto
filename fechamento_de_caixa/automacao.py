import os
import re
import time
import subprocess
from tkinter import Tk, filedialog
import openpyxl
from pywinauto import Application
import pyautogui
from PIL import Image
import pytesseract
import pandas as pd
import cv2
import numpy as np

# --- CONFIGURAÇÕES DO ROBÔ (ALTERE SE PRECISAR) ---
QUANTIDADE_DIGITOS = 4       # Quantos últimos números da autorização conferir
VELOCIDADE_DIGITACAO = 0.08  # Tempo de espera entre cada tecla digitada (segundos)

def localizar_e_clicar_imagem(nome_imagem, cliques=1, arrasto_x=0, arrasto_y=0):
    """Localiza uma imagem de referência na tela e retorna o ponto central encontrado."""
    diretorio_projeto = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    caminho_real_imagem = os.path.join(diretorio_projeto, nome_imagem)
    
    print_tela = pyautogui.screenshot()
    imagem_tela = cv2.cvtColor(np.array(print_tela), cv2.COLOR_RGB2BGR)
    imagem_alvo = cv2.imread(caminho_real_imagem)
    
    if imagem_alvo is None:
        return None
        
    resultado = cv2.matchTemplate(imagem_tela, imagem_alvo, cv2.TM_CCOEFF_NORMED)
    _, max_val, _, max_loc = cv2.minMaxLoc(resultado)
    
    if max_val >= 0.8:
        h, w = imagem_alvo.shape[:2]
        centro_x = max_loc[0] + (w // 2) + arrasto_x
        centro_y = max_loc[1] + (h // 2) + arrasto_y
        
        pyautogui.moveTo(centro_x, centro_y, duration=0.4)
        if cliques > 0:
            pyautogui.click(clicks=cliques, interval=0.2)
        return (centro_x, centro_y)
    return None

def extrair_dados_do_nome(nome_arquivo):
    nome_sem_extensao, _ = os.path.splitext(nome_arquivo)
    partes = nome_sem_extensao.split(" ")
    cartao = "Goodcard" if "Goodcard" in nome_sem_extensao else "Não identificado"
    posto = "Não identificado"
    if "Centro Automotivo" in nome_sem_extensao:
        posto = "Centro Automotivo"
    elif "Arquipelago" in nome_sem_extensao:
        posto = "Posto Archipelago"
    elif "Mareli" in nome_sem_extensao:
        posto = "Posto Mareli"
    data_lancamento = partes[-1] if partes else "Não identificada"
    return cartao, posto, data_lancamento


def ler_texto_celula(x_centro, y_centro, largura=100, altura=22):
    x1 = max(0, x_centro - largura // 2)
    y1 = max(0, y_centro - altura // 2)
    imagem = pyautogui.screenshot(region=(x1, y1, largura, altura))
    texto = pytesseract.image_to_string(np.array(imagem), config='--psm 7 --oem 3')
    return str(texto).strip()


def extrair_valor_ocr(texto):
    texto = str(texto or "")
    if not texto:
        return ""

    padroes = re.findall(r"\d{1,3}(?:[.\s]\d{3})*(?:[.,]\d{2})?|\d{1,3}[.,]\d{2}", texto)
    if not padroes:
        return ""

    ultimo = padroes[-1].replace(" ", "")
    if ultimo.count(',') == 0 and ultimo.count('.') == 0:
        return ultimo
    if ',' in ultimo and '.' in ultimo:
        ultimo = ultimo.replace('.', '').replace(',', '.')
    elif ',' in ultimo:
        ultimo = ultimo.replace('.', '').replace(',', '.')
    return ultimo


def processar_planilha_seguro(caminho_arquivo):
    print("[PLANILHA] Verificando e estruturando colunas da planilha selecionada...")
    if caminho_arquivo.endswith('.xls'):
        pasta_destino = os.path.dirname(caminho_arquivo)
        nome_base, _ = os.path.splitext(caminho_arquivo)
        caminho_trabalho = nome_base + ".xlsx"
        caminhos_libreoffice = [
            r"C:\Program Files\LibreOffice\program\scalc.exe",
            r"C:\Program Files (x86)\LibreOffice\program\scalc.exe"
        ]
        scalc_path = next((c for c in caminhos_libreoffice if os.path.exists(c)), None)
        if scalc_path:
            comando = f'"{scalc_path}" --headless --convert-to xlsx "{caminho_arquivo}" --outdir "{pasta_destino}"'
            subprocess.run(comando, shell=True, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            try: os.remove(caminho_arquivo) 
            except: pass
        else:
            return False
    else:
        caminho_trabalho = caminho_arquivo
    try:
        wb = openpyxl.load_workbook(caminho_trabalho)
        ws = wb.active
        linha_cnpj = None
        for row in range(1, 20):
            val_a = str(ws.cell(row=row, column=1).value).strip()
            val_b = str(ws.cell(row=row, column=2).value).strip()
            if val_a == "CNPJ" or val_b == "CNPJ":
                linha_cnpj = row
                break
        if not linha_cnpj:
            linha_cnpj = 6
        ws.insert_cols(1)
        ws.cell(row=linha_cnpj, column=1).value = "Conferido"
        pasta = os.path.dirname(caminho_trabalho)
        nome_completo = os.path.basename(caminho_trabalho)
        nome_sem_ext, _ = os.path.splitext(nome_completo)
        if nome_sem_ext.endswith("_conferido"):
            nome_sem_ext = nome_sem_ext.replace("_conferido", "")
        caminho_final = os.path.join(pasta, f"{nome_sem_ext}_conferido.xlsx")
        wb.save(caminho_trabalho)
        wb.close()
        if os.path.exists(caminho_trabalho):
            if os.path.exists(caminho_final):
                os.remove(caminho_final)
            os.rename(caminho_trabalho, caminho_final)
        print(f"[SUCESSO] Planilha editada e salva como: {os.path.basename(caminho_final)}")
        return caminho_final
    except Exception as e:
        print(f"[ERRO] Falha crítica ao manipular as células da planilha: {e}")
        return False

def selecionar_e_processar_projeto():
    print("Aguardando seleção do arquivo pelo usuário...")
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo de Fechamento de Caixa",
        filetypes=[("Arquivos de Fechamento", "*.xls *.xlsx")]
    )
    if not caminho_arquivo:
        return None
    caminho_arquivo = os.path.abspath(caminho_arquivo)
    nome_arquivo = os.path.basename(caminho_arquivo)
    cartao, posto, data_lancamento = extrair_dados_do_nome(nome_arquivo)
    print(f"\n[DADOS CAPTURADOS DO NOME]")
    print(f"💳 Cartão: {cartao}")
    print(f"⛽ Posto: {posto}")
    print(f"📅 Data: {data_lancamento}\n")
    caminho_salvo = processar_planilha_seguro(caminho_arquivo)
    return {"cartao": cartao, "posto": posto, "data": data_lancamento, "caminho_excel": caminho_salvo}

def abrir_e_logar_webposto():
    dados_caixa = selecionar_e_processar_projeto()
    if not dados_caixa or not dados_caixa.get("caminho_excel"):
        return
    print("\n[SISTEMA] Iniciando a automação do sistema webPosto...")
    caminho_programa = r"C:\Quality\web\QualityPosto.exe"
    try:
        app = Application(backend="win32").start(caminho_programa)
        time.sleep(5)
        janela_login = app.top_window()
        janela_login.set_focus()
        janela_login.click_input()
        janela_login.type_keys("^a{BACKSPACE}GESSICAMARELI", set_foreground=True)
        janela_login.type_keys("{TAB}")
        janela_login.type_keys("12345")
        janela_login.type_keys("{ENTER}")
        
        diretorio_projeto = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        pytesseract.pytesseract.tesseract_cmd = os.path.join(diretorio_projeto, "Tesseract-OCR", "tesseract.exe")
        time.sleep(4)
        janela_principal = app.top_window()
        janela_principal.set_focus()
        
        posto_detectado = dados_caixa['posto']
        pulos_seta = 0
        if posto_detectado == "Centro Automotivo": pulos_seta = 0
        elif posto_detectado == "Posto Mareli": pulos_seta = 1
        elif posto_detectado == "Posto Archipelago": pulos_seta = 2
            
        pyautogui.click(x=janela_principal.rectangle().left + 200, y=janela_principal.rectangle().top + 180)
        time.sleep(0.5)
        for _ in range(pulos_seta):
            pyautogui.press('down')
            time.sleep(0.3)
        pyautogui.press('enter')
        time.sleep(8)
        
        janela_principal = app.top_window()
        janela_principal.set_focus()
        try:
            rect = janela_principal.rectangle()
            for _ in range(3):
                pyautogui.click(x=rect.left + 25, y=rect.bottom - 20)
                time.sleep(0.8)
        except: pass
        pyautogui.press('esc')
        time.sleep(1.5)
        
        if not localizar_e_clicar_imagem('aba_financeiro_menu.png', cliques=1):
            return
        time.sleep(2)
        if not localizar_e_clicar_imagem('fechamento_de_caixa.png', cliques=1):
            return
        time.sleep(4) 
        
        data_para_preencher = dados_caixa['data'].replace("-", "").replace("/", "")
        ponto_inicio = localizar_e_clicar_imagem('campo_data.png', cliques=2)
        if ponto_inicio:
            x_ini, y_ini = ponto_inicio
            pyautogui.hotkey('ctrl', 'a')
            pyautogui.press('backspace')
            pyautogui.write(data_para_preencher, interval=0.05)
            pyautogui.press('esc')
            time.sleep(0.5)
            pyautogui.moveTo(x_ini, y_ini + 35, duration=0.4)
            pyautogui.click(clicks=2, interval=0.2)
            time.sleep(0.3)
            pyautogui.hotkey('ctrl', 'a')
            pyautogui.press('backspace')
            pyautogui.write(data_para_preencher, interval=0.05)
            pyautogui.press('esc')
            time.sleep(0.5)
        else:
            return

        ponto_consultar = localizar_e_clicar_imagem('botao_consultar.png')
        if not ponto_consultar:
            return
        y_btn = ponto_consultar[1]
        time.sleep(4)

        if not localizar_e_clicar_imagem('texto_1turno.png', cliques=2):
            return
        time.sleep(5)
        if not localizar_e_clicar_imagem('aba_cartao.png'):
            return
        time.sleep(2)

        def executar_filtro_goodcard(tentativa=1):
            print(f"\n[FILTRO] Analisando a lista de administradoras (Página {tentativa})...")
            caminho_admin = os.path.join(diretorio_projeto, 'texto_administradora.png')
            print_tela = pyautogui.screenshot()
            img_tela = cv2.cvtColor(np.array(print_tela), cv2.COLOR_RGB2BGR)
            img_alvo = cv2.imread(caminho_admin)
            if img_alvo is not None:
                res = cv2.matchTemplate(img_tela, img_alvo, cv2.TM_CCOEFF_NORMED)
                _, max_val, _, max_loc = cv2.minMaxLoc(res)
                if max_val >= 0.8:
                    h, w = img_alvo.shape[:2]
                    pyautogui.moveTo(max_loc[0] + w + 15, max_loc[1] + (h // 2), duration=0.4)
                    time.sleep(0.6)
            
            if localizar_e_clicar_imagem('filtro_administradora.png', cliques=1):
                time.sleep(1.5)
                caminho_goodcard = os.path.join(diretorio_projeto, 'opcao_goodcard.png')
                print_lista = pyautogui.screenshot()
                img_lista = cv2.cvtColor(np.array(print_lista), cv2.COLOR_RGB2BGR)
                img_good = cv2.imread(caminho_goodcard)
                
                encontrou_goodcard = False
                if img_good is not None:
                    res_good = cv2.matchTemplate(img_lista, img_good, cv2.TM_CCOEFF_NORMED)
                    _, max_val_good, _, max_loc_good = cv2.minMaxLoc(res_good)
                    if max_val_good >= 0.8:
                        h_g, w_g = img_good.shape[:2]
                        pyautogui.moveTo(max_loc_good[0] + (w_g // 2), max_loc_good[1] + (h_g // 2), duration=0.4)
                        pyautogui.click()
                        print(f"[🎯 FILTRO] GOODCARD aplicado com sucesso na Página {tentativa}!")
                        encontrou_goodcard = True
                        time.sleep(2)
                        executar_conferencia_valores(dados_caixa['caminho_excel'], y_btn)
                
                if not encontrou_goodcard:
                    print(f"[AVISO] GOODCARD não está na Página {tentativa}. Avançando...")
                    pyautogui.press('esc')
                    time.sleep(1)
                    if localizar_e_clicar_imagem('botao_avancar_caixa.png', cliques=1):
                        time.sleep(4)
                        executar_filtro_goodcard(tentativa = tentativa + 1)

        executar_filtro_goodcard()
    except Exception as e:
        print(f"[ERRO] Falha geral: {e}")

def executar_conferencia_valores(caminho_excel, y_btn):
    print("\n[CONFERÊNCIA] Iniciando a varredura dinâmica linha por linha (Suporta Scroll)...")
    try:
        df_cru = pd.read_excel(caminho_excel, header=None)
        linha_cabecalho_real = 0
        coluna_autorizacao = None
        coluna_valor_bruto = None
        
        for index_linha, linha in df_cru.head(15).iterrows():
            valores_linha = [str(celula).strip().lower() for celula in linha.values]
            if any("autoriza" in cel or "aut" in cel for cel in valores_linha) and any("valor" in cel or "bruto" in cel for cel in valores_linha):
                linha_cabecalho_real = index_linha
                break
                
        wb_conferencia = openpyxl.load_workbook(caminho_excel)
        ws_conferencia = wb_conferencia.active
        
        df_planilha = pd.read_excel(caminho_excel, skiprows=linha_cabecalho_real)
        df_planilha.columns = [str(c).strip() for c in df_planilha.columns]
        
        coluna_ok_index = 1
        for col_idx in range(1, ws_conferencia.max_column + 1):
            if str(ws_conferencia.cell(row=linha_cabecalho_real + 1, column=col_idx).value).strip().lower() == "conferido":
                coluna_ok_index = col_idx
                break
        
        for col in df_planilha.columns:
            col_limpa = str(col).strip().lower()
            if "autoriza" in col_limpa or "aut" in col_limpa: coluna_autorizacao = col
            elif "bruto" in col_limpa or col_limpa == "valor bruto": coluna_valor_bruto = col

        if not coluna_valor_bruto:
            for col in df_planilha.columns:
                if "valor" in str(col).lower() and "reembolso" not in str(col).lower():
                    coluna_valor_bruto = col
                    break

        if not coluna_autorizacao or not coluna_valor_bruto:
            print("[ERRO] Não consegui mapear as colunas da planilha.")
            return
            
        print(f"[SISTEMA] Colunas validadas -> Autorização: '{coluna_autorizacao}' | Valor Correto: '{coluna_valor_bruto}'")
        
        pos_col_autorizacao = localizar_e_clicar_imagem('titulo_autorizacao.png', cliques=0)
        pos_col_valor = localizar_e_clicar_imagem('titulo_valor.png', cliques=0)
        pos_col_conferido = localizar_e_clicar_imagem('titulo_conferido.png', cliques=0)
        
        if not pos_col_autorizacao or not pos_col_valor or not pos_col_conferido:
            print("[ERRO] Cabeçalhos de coluna não identificados no webPosto.")
            return
            
        x_col_autorizacao, y_grid_header = pos_col_autorizacao if isinstance(pos_col_autorizacao, tuple) else (pos_col_autorizacao, y_btn)
        x_col_valor, _ = pos_col_valor if isinstance(pos_col_valor, tuple) else (pos_col_valor, 0)
        x_col_conferido, _ = pos_col_conferido if isinstance(pos_col_conferido, tuple) else (pos_col_conferido, 0)
        
        lista_conferencia = []
        for index, row in df_planilha.iterrows():
            aut_c = str(row[coluna_autorizacao]).strip().split('.')
            aut_limpa = aut_c[0] if aut_c else ""
            if aut_limpa == "nan" or aut_limpa == "" or "total" in aut_limpa.lower(): continue
            if aut_limpa.endswith('.0'): aut_limpa = aut_limpa[:-2]
            
            aut_f = aut_limpa[-QUANTIDADE_DIGITOS:] if len(aut_limpa) >= QUANTIDADE_DIGITOS else aut_limpa
            val_b = str(row[coluna_valor_bruto]).replace("R$", "").replace(" ", "").strip()
            try: val_f = float(val_b.replace(".", "").replace(",", "."))
            except: continue
            lista_conferencia.append({"final_aut": aut_f, "valor_bruto_str": val_b, "valor_bruto_float": val_f, "linha_excel_real": index + linha_cabecalho_real + 2})

        total_planilha = len(lista_conferencia)
        print(f"[SISTEMA] Total de registros localizados para conferência: {total_planilha}")
        if total_planilha == 0: return

        print("[ROBÔ] Ativando foco na primeira linha de dados do webPosto...")
        pyautogui.moveTo(x_col_autorizacao, y_grid_header + 25, duration=0.4)
        pyautogui.click()
        time.sleep(0.5)
        
        import pyperclip

        autorizacoes_processadas = []
        linhas_repetidas_seguidas = 0
        conferencias_realizadas_com_sucesso = 0
        ultima_selecao_invalida = ""
        
        # Guarda a altura inicial aproximada da primeira linha do grid de dados
        y_linha_dinamica = y_grid_header + 25

        # Inicia a varredura dinâmica linha por linha
        while linhas_repetidas_seguidas < 3:
            if conferencias_realizadas_com_sucesso >= total_planilha:
                break

            # Leitura OCR da própria célula da autorização, sem copiar a linha inteira do grid.
            print("[SISTEMA] Capturando número real da Autorização via OCR da célula...")
            texto_capturado = ler_texto_celula(x_col_autorizacao + 20, y_linha_dinamica, largura=80, altura=18)
            texto_capturado = texto_capturado.replace('\n', ' ').strip()

            if not texto_capturado or 'R$' in texto_capturado or 'Cliente' in texto_capturado or 'Conferido' in texto_capturado or any(ch.isalpha() for ch in texto_capturado):
                print(f"[AVISO] Seleção descartada como valor/coluna errada: '{texto_capturado}'")

                if texto_capturado == ultima_selecao_invalida:
                    linhas_repetidas_seguidas += 1
                    if linhas_repetidas_seguidas >= 3:
                        print("[SISTEMA] A mesma seleção inválida foi repetida 3 vezes. Encerrando varredura para evitar loop infinito.")
                        break
                else:
                    linhas_repetidas_seguidas = 0
                    ultima_selecao_invalida = texto_capturado

                pyautogui.click(x_col_autorizacao, y_linha_dinamica)
                time.sleep(0.1)
                pyautogui.press('down')
                time.sleep(0.5)
                y_linha_dinamica += 20
                continue
            else:
                ultima_selecao_invalida = ""
                linhas_repetidas_seguidas = 0

            aut_posto_completa = "".join(c for c in texto_capturado if c.isdigit())
            aut_posto_final = aut_posto_completa[-QUANTIDADE_DIGITOS:] if len(aut_posto_completa) >= QUANTIDADE_DIGITOS else aut_posto_completa

            # Se a tabela chegou ao fim ou pegou uma linha vazia, encerra limpamente
            if aut_posto_final in ["000", "0000", ""] or not aut_posto_final:
                print("\n[SISTEMA] Fim dos registros válidos da tabela atingido!")
                break
                
            # Evita loops infinitos caso a mesma linha continue sendo lida repetidamente.
            if autorizacoes_processadas and aut_posto_final == autorizacoes_processadas[-1]:
                linhas_repetidas_seguidas += 1
                if linhas_repetidas_seguidas >= 5:
                    print("[SISTEMA] Repetição da mesma linha detectada. Encerrando varredura.")
                    break
                pyautogui.press('down')
                time.sleep(0.5)
                y_linha_dinamica += 20
                continue
            else:
                linhas_repetidas_seguidas = 0
                
            print(f"\n[VARREDURA] Linha atual detectada: Autorização real final '{aut_posto_final}'")
            autorizacoes_processadas.append(aut_posto_final)
            
            # Busca se essa autorização exata consta na nossa lista do Excel
            item_correspondente = next((item for item in lista_conferencia if item["final_aut"] == aut_posto_final), None)
            
            if item_correspondente:
                val_excel_bruto = str(item_correspondente["valor_bruto_str"]).strip()
                num_linha_excel = item_correspondente["linha_excel_real"]
                
                # Garante a formatação de centavos com vírgula padrão brasileiro
                if "," not in val_excel_bruto and "." not in val_excel_bruto:
                    val_excel_bruto = f"{val_excel_bruto},00"
                else:
                    val_excel_bruto = val_excel_bruto.replace(".", ",")

                # Leitura OCR do valor da própria célula para evitar pegar a linha inteira do grid.
                print("[SISTEMA] Capturando valor real do programa via OCR da célula...")
                valor_posto_texto = ler_texto_celula(x_col_valor + 55, y_linha_dinamica, largura=90, altura=18)
                valor_posto_texto = extrair_valor_ocr(valor_posto_texto)
                valor_posto_texto = str(valor_posto_texto).replace("R$", "").replace(" ", "").replace("\n", "").strip()
                
                # --- ETAPA 3: CONFERÊNCIA CIRÚRGICA DE STRINGS ---
                if valor_posto_texto == val_excel_bruto or val_excel_bruto in valor_posto_texto:
                    print(f"Valor Planilha | Valor Programa: {val_excel_bruto} | {valor_posto_texto}")
                    pyautogui.moveTo(x_col_conferido, y_linha_dinamica, duration=0.2)
                    pyautogui.click()
                    time.sleep(0.2)
                    ws_conferencia.cell(row=num_linha_excel, column=coluna_ok_index).value = "OK"
                else:
                    print(f"Valor Planilha | Valor Programa: {val_excel_bruto} | {valor_posto_texto}")
                    print(f"Alterando o valor {valor_posto_texto} para o valor {val_excel_bruto}")
                    
                    # Dá o duplo clique na coluna Valor na altura exata da linha selecionada
                    pyautogui.moveTo(x_col_valor, y_linha_dinamica, duration=0.3)
                    pyautogui.click(clicks=2, interval=0.2)
                    time.sleep(0.4)
                    
                    pyautogui.hotkey('ctrl', 'a')
                    pyautogui.press('backspace')
                    time.sleep(0.1)
                    
                    for caractere in val_excel_bruto:
                        pyautogui.write(caractere)
                        time.sleep(VELOCIDADE_DIGITACAO)
                    pyautogui.press('enter')
                    time.sleep(0.6)
                    
                    pyautogui.moveTo(x_col_conferido, y_linha_dinamica, duration=0.2)
                    pyautogui.click()
                    time.sleep(0.2)
                    ws_conferencia.cell(row=num_linha_excel, column=coluna_ok_index).value = "OK"
                
                conferencias_realizadas_com_sucesso += 1
            else:
                print(f"[INFO] Autorização {aut_posto_final} do posto não consta na planilha selecionada.")
                
            # Clica no centro da linha para fixar o foco antes de descer pelo teclado
            pyautogui.click(x_col_autorizacao, y_linha_dinamica)
            time.sleep(0.1)
            pyautogui.press('down')
            time.sleep(0.5)
            y_linha_dinamica += 20
            
        wb_conferencia.save(caminho_excel)
        wb_conferencia.close()
        print(f"\n[💾 PLANILHA] Alterações de 'OK' gravadas e salvas com sucesso em: {os.path.basename(caminho_excel)}")
        print(f"[🎯 CONFERÊNCIA FINALIZADA] Total realizado: {conferencias_realizadas_com_sucesso} de {total_planilha} registros.")
    except Exception as e:
        print(f"[ERRO CRÍTICO] Falha no bloco de conferencia: {e}")

if __name__ == "__main__":
    abrir_e_logar_webposto()
